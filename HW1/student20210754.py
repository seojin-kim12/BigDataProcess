
#!/usr/bin/python3
from openpyxl import load_workbook

try:
    workbook = load_workbook("student.xlsx")

    sheet = workbook.active

    for row_number in range(2, 76):
        midterm = sheet.cell(row=row_number, column=3).value
        final = sheet.cell(row=row_number, column=4).value
        hw = sheet.cell(row=row_number, column=5).value
        attendance = sheet.cell(row=row_number, column=6).value

        total = midterm * 0.3 + final * 0.35 + hw * 0.34 + attendance * 0.01

        sheet.cell(row=row_number, column=7, value=total)

    # 학점 주는 코드
    for row_num in range(2, 76):
        total = sheet.cell(row=row_num, column=7).value

        if total < 40:
            grade = "F"
        else:
            # total 값으로 학생들을 정렬
            sorted_students = sorted(range(2, 76), key=lambda x: sheet.cell(row=x, column=7).value, reverse=True)

            # 상위 30% 학생에게 A 부여
            a_count = int(0.3 * len(sorted_students))
            for i in range(a_count):
                sheet.cell(row=sorted_students[i], column=8, value="A")

            # A 학점 받은 학생들을 제외하고 나머지에게 B 부여
            b_count = int(0.7 * len(sorted_students)) - a_count
            for i in range(a_count, a_count + b_count):
                sheet.cell(row=sorted_students[i], column=8, value="B")

            # A와 B 학점 받은 학생들을 제외하고 나머지에게 C 부여
            c_count = len(sorted_students) - (a_count + b_count)
            for i in range(a_count + b_count, len(sorted_students)):
                sheet.cell(row=sorted_students[i], column=8, value="C")

            # A 학점 내에서 상위 50%에게 A+ 부여
            a_students = [row for row in sorted_students[:a_count] if sheet.cell(row=row, column=8).value == "A"]
            a_plus_count = int(0.5 * len(a_students))
            for i in range(a_plus_count):
                sheet.cell(row=a_students[i], column=8, value="A+")

            # B 학점 내에서 상위 50%에게 B+ 부여
            b_students = [row for row in sorted_students[a_count:a_count + b_count] if sheet.cell(row=row, column=8).value == "B"]
            b_plus_count = int(0.5 * len(b_students))
            for i in range(b_plus_count):
                sheet.cell(row=b_students[i], column=8, value="B+")

            # C 학점 내에서 상위 50%에게 C+ 부여
            c_students = [row for row in sorted_students[a_count + b_count:] if sheet.cell(row=row, column=8).value == "C"]
            c_plus_count = int(0.5 * len(c_students))
            for i in range(c_plus_count):
                sheet.cell(row=c_students[i], column=8, value="C+")

            # 총점이 40점 미만인 학생에게 F 부여
            for row in range(2, 76):
                total = sheet.cell(row=row, column=7).value
                if total < 40:
                    sheet.cell(row=row, column=8, value="F")

    workbook.save("student.xlsx")
    print("실행 완료")
except Exception as e:
    print(f"스크립트 실행 중 오류 발생: {str(e)}")